from flask import Flask, request, jsonify, render_template
from flask_cors import CORS
from api_stocks import get_stock_data
from api_crypto import get_daily_crypto_data, get_exchange_rate
from chatbot import chatbot_response
import os
from PIL import Image
import pytesseract
import PyPDF2
import docx
import openpyxl

app = Flask(__name__)
CORS(app)  # Enable CORS if needed

# Set the upload folder path
app.config['UPLOAD_FOLDER'] = os.path.join(app.root_path, 'uploads')

# Ensure the uploads directory exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

@app.route('/')
def home():
    return render_template('index.html')  # Serve the HTML file

@app.route('/chatbot', methods=['POST'])
def chatbot():
    data = request.json
    question = data.get("question", "")
    response = chatbot_response(question)  # Call your function to get the response
    return jsonify({"response": response})

# Endpoint for file upload
@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    # Validate the file extension
    allowed_extensions = {'pdf', 'docx', 'xlsx', 'png', 'jpg', 'jpeg'}
    file_extension = file.filename.rsplit('.', 1)[1].lower()

    if file_extension not in allowed_extensions:
        return jsonify({'error': 'File type not allowed'}), 400

    # Save the file
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(file_path)

    # Process the file and extract text
    extracted_text = ''
    if file_extension == 'pdf':
        extracted_text = extract_text_from_pdf(file_path)
    elif file_extension == 'docx':
        extracted_text = extract_text_from_docx(file_path)
    elif file_extension == 'xlsx':
        extracted_text = extract_text_from_xlsx(file_path)
    elif file_extension in {'png', 'jpg', 'jpeg'}:
        extracted_text = extract_text_from_image(file_path)

    return jsonify({'message': 'File uploaded successfully', 'filename': file.filename, 'extracted_text': extracted_text}), 200

def extract_text_from_pdf(file_path):
    text = ''
    with open(file_path, 'rb') as f:
        reader = PyPDF2.PdfReader(f)
        for page in reader.pages:
            text += page.extract_text() or ''
    return text

def extract_text_from_docx(file_path):
    text = ''
    doc = docx.Document(file_path)
    for paragraph in doc.paragraphs:
        text += paragraph.text + '\n'
    return text

def extract_text_from_xlsx(file_path):
    text = ''
    workbook = openpyxl.load_workbook(file_path)
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        for row in ws.iter_rows(values_only=True):
            text += ' '.join([str(cell) for cell in row if cell is not None]) + '\n'
    return text

def extract_text_from_image(file_path):
    text = ''
    image = Image.open(file_path)
    text = pytesseract.image_to_string(image)
    return text

    
# Endpoint for stock data
@app.route('/api/stock/<string:symbol>', methods=['GET'])
def stock_data(symbol):
    try:
        data, sma_data = get_stock_data(symbol)
        if data is None or sma_data is None:
            return jsonify({'error': 'Failed to fetch stock data.'}), 500
        
        response_data = {
            'stock_data': data.reset_index().to_dict(orient='records'),
            'sma_data': sma_data.reset_index().to_dict(orient='records')
        }
        return jsonify(response_data), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Endpoint for cryptocurrency data
@app.route('/api/crypto/<string:symbol>/<string:market>', methods=['GET'])
def crypto_data(symbol, market):
    try:
        daily_data = get_daily_crypto_data(symbol, market)
        if daily_data is None:
            return jsonify({'error': 'Failed to fetch cryptocurrency data.'}), 500
        
        exchange_rate = get_exchange_rate(symbol, market)
        response_data = {
            'daily_data': daily_data.reset_index().to_dict(orient='records'),
            'exchange_rate': exchange_rate
        }
        return jsonify(response_data), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)  # This runs the server in debug mode
